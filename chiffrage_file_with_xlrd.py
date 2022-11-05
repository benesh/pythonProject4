import openpyxl as op

path = r"C:\Users\cobcouli\OneDrive - Capgemini\Documents\Python_ressources\Feuille de calcul dans Exercice python_Jouer_avec_des_donnees.xlsx"

def chiffrage_xlrd():
    book = op.load_workbook(path)
    ws = book['Gestion_des_Rapporteurs']

    """qustion num 1 pour chaque scenario avoir les nombre de cas de test"""
    num_col_scenario = 3
    char = op.utils.get_column_letter(num_col_scenario)

    """suppression de ligne vide"""
    num_col_for_suppression = 3
    char_for_deletion = getcolumn(num_col_for_suppression)
    maxrow= ws.max_row
    compteur = 0
    #delete row null
    for row in ws:
        delete_raw_in_sheet(ws,row)
        compteur += 1

    maxrow = ws.max_row
    print(ws.max_row)

    """ initialisation """
    #initaliser le dictionnaire conteneur des scénario et cas de test associé
    dictionary_of_scenario = {}

    """1 Parcours de la colonne et détection des different scenraio et le nombre de leurs cas de test """
    for row in range(2, ws.max_row):
        if (ws[char + str(row)].value in dictionary_of_scenario.keys()):
            if ws[getcolumn(num_col_scenario + 1) + str(row)].value not in dictionary_of_scenario.get(ws[char + str(row)].value):
                listTemp = dictionary_of_scenario.get(ws[char + str(row)].value)
                listTemp.append(ws[getcolumn(num_col_scenario + 1) + str(row)].value)
                dictionary_of_scenario[ws[char + str(row)].value] = listTemp
        else:
            dictionary_of_scenario[ws[char + str(row)].value] = [ws[getcolumn(num_col_scenario + 1) + str(row)].value]

    """ dictionnaire contenant tous les infos et affichage de la réponse"""
    del dictionary_of_scenario[None] # suppression des None car le néttoyage pas propre

    """ 1 Pour chaque scénario de test, donner le nombre de cas de test associé"""
    sheet_nbre_cas = book.create_sheet("Scenario_avec_nbr_cas_de_test")
    heads = ['Scenario de Test','Nombre de Cas de test','only fans']
    sheet_nbre_cas.append(heads)
    for key in dictionary_of_scenario:
        sheet_nbre_cas.append([key,len(dictionary_of_scenario.get(key))])

    """Donner pour tout cas de test leur nombre de step"""
    sheet_nbr_detape = book.create_sheet("Nombre de Step par cas")
    sheet_nbr_detape.append(['Cas de test','Nombre d\'étape'])

    dictionary_of_cas_de_test = {}
    num_col_cas_de_test = 4 # colonne cas de test
    char_cas_de_test = getcolumn(num_col_cas_de_test)
    """boucle de dictionnaire"""
    for row in range(2, ws.max_row):
        if (ws[char_cas_de_test + str(row)].value in dictionary_of_cas_de_test.keys()):
            if ws[getcolumn(num_col_cas_de_test + 1) + str(row)].value not in dictionary_of_cas_de_test.get(ws[char_cas_de_test + str(row)].value):
                listTemp = dictionary_of_cas_de_test.get(ws[char_cas_de_test + str(row)].value)
                listTemp.append(ws[getcolumn(num_col_cas_de_test + 1) + str(row)].value)
                dictionary_of_cas_de_test[ws[char_cas_de_test + str(row)].value] = listTemp
        else:
            dictionary_of_cas_de_test[ws[char_cas_de_test + str(row)].value] = [ws[getcolumn(num_col_cas_de_test + 1) + str(row)].value]

    del dictionary_of_cas_de_test[None]

    for key in dictionary_of_cas_de_test:
        sheet_nbr_detape.append([key,len(dictionary_of_cas_de_test.get(key))])

    """ stat a donner """
    sheet_stat_a_donner = book.create_sheet("Stat a donner")

    """Nombre moyen de step par cas de test"""
    nbr_total_cas_de_test = len(dictionary_of_cas_de_test.keys())
    items = dictionary_of_cas_de_test.items()
    nbre_total_step = 0
    for item in items:
        nbre_total_step += len(item)

    nbr_moyen_de_step_par_cas_de_test =nbre_total_step/nbr_total_cas_de_test
    sheet_stat_a_donner.append(['Nombre Moyen par Cas de Test',nbr_moyen_de_step_par_cas_de_test])

    """nombre de cas de test avec des étapes =< 3 """
    nbr_cas_de_test_avec_step_inf_3 = 0
    for key in  dictionary_of_cas_de_test:
        if (len(dictionary_of_cas_de_test.get(key))<= 3):
            nbr_cas_de_test_avec_step_inf_3 += 1
    sheet_stat_a_donner.append(['Nomre de cas dse test <= 3',nbr_cas_de_test_avec_step_inf_3])

    """nombre de cas de test entre 4 et 7 étape"""
    nbr_cas_de_test_avec_step_entre_4_7 = 0
    for key in  dictionary_of_cas_de_test:
        if (len(dictionary_of_cas_de_test.get(key))>= 4 & len(dictionary_of_cas_de_test.get(key))<= 7 ):
            nbr_cas_de_test_avec_step_entre_4_7 += 1
    sheet_stat_a_donner.append(['Nombre de cas de test entre 4 et 7',nbr_cas_de_test_avec_step_entre_4_7])

    """nombre de cas de test ayant plus ou égal de 8 étape"""
    nbr_cas_de_test_avec_step_sup_8 = 0
    for key in dictionary_of_cas_de_test:
        if (len(dictionary_of_cas_de_test.get(key)) >= 8):
            nbr_cas_de_test_avec_step_sup_8 += 1
    sheet_stat_a_donner.append(['Nombre de cas de test sup ou égal a 8',nbr_cas_de_test_avec_step_sup_8])


    book.save("tempexcell.xlsx")

def getcolumn(numbre):
    return op.utils.get_column_letter(numbre)

def delete_raw_in_sheet(sheet,row):
    for cell in row:
        if cell.value != None:
            return
    sheet.delete_rows(row[0].row)

